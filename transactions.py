import os
from datetime import datetime

from aiogram import Router, F
from aiogram.types import Message, FSInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import database as db

router = Router()


@router.message(F.text == "📊 Umumiy hisobot")
async def general_report(message: Message):
    income, expense = await db.get_totals()
    balance = income - expense
    total_debt = await db.get_total_debt()
    products = await db.get_all_products()
    total_stock_value = sum(p["price"] * p["quantity"] for p in products)

    text = (
        "📊 <b>Umumiy hisobot</b>\n\n"
        f"💰 Kirim: {income:.0f} so'm\n"
        f"💸 Chiqim: {expense:.0f} so'm\n"
        f"📈 Balans: <b>{balance:.0f} so'm</b>\n\n"
        f"📦 Skladdagi mahsulotlar soni: {len(products)} xil\n"
        f"📦 Sklad qiymati: {total_stock_value:.0f} so'm\n\n"
        f"📒 Umumiy qarzdorlik: {total_debt:.0f} so'm"
    )
    await message.answer(text, parse_mode="HTML")


@router.message(F.text == "📥 Excel yuklab olish")
async def export_excel(message: Message):
    products = await db.get_all_products()
    transactions = await db.get_transactions(limit=1000)
    debts = await db.get_debts(only_unpaid=False)

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # ---- Sklad ----
    ws1 = wb.active
    ws1.title = "Sklad"
    ws1.append(["ID", "Nomi", "Narxi", "Miqdori", "Jami qiymat", "Qo'shilgan sana"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
    for p in products:
        ws1.append([
            p["id"], p["name"], p["price"], p["quantity"],
            p["price"] * p["quantity"], p["created_at"][:10]
        ])

    # ---- Kirim/Chiqim ----
    ws2 = wb.create_sheet("Kirim-Chiqim")
    ws2.append(["ID", "Turi", "Summasi", "Izoh", "Sana"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for t in transactions:
        label = "Kirim" if t["type"] == "income" else "Chiqim"
        ws2.append([t["id"], label, t["amount"], t["description"], t["created_at"][:16]])

    # ---- Qarz daftar ----
    ws3 = wb.create_sheet("Qarz daftar")
    ws3.append(["ID", "Mijoz", "Telefon", "Summasi", "Izoh", "Holati", "Sana"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for d in debts:
        status = "To'landi" if d["is_paid"] else "Qarzda"
        ws3.append([
            d["id"], d["customer_name"], d["phone"], d["amount"],
            d["description"], status, d["created_at"][:10]
        ])

    # Ustun kengligini avtomoslashtirish
    for ws in (ws1, ws2, ws3):
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 3

    filename = f"hisobot_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)

    await message.answer_document(FSInputFile(filepath, filename=filename))
    os.remove(filepath)
