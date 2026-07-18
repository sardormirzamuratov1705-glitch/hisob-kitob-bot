import os
import shutil
from datetime import datetime

import aiosqlite
from aiogram import Router, F
from aiogram.types import Message, FSInputFile, CallbackQuery
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import config
import database as db
import keyboards as kb
from access_control import get_shop_id, is_admin

router = Router()


class RestoreDB(StatesGroup):
    waiting_file = State()


async def _require_shop(message: Message):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await message.answer("Bu bo'lim faqat do'kon egalari uchun.")
    return shop_id


@router.message(F.text == "📊 Umumiy hisobot")
async def general_report(message: Message):
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    income, expense = await db.get_totals(shop_id)
    balance = income - expense
    total_debt = await db.get_total_debt(shop_id)
    products = await db.get_all_products(shop_id)
    total_stock_value = sum(p["price"] * p["quantity"] for p in products)
    payment_totals = await db.get_payment_method_totals(shop_id, "income")

    text = (
        "📊 <b>Umumiy hisobot</b>\n\n"
        f"💰 Kirim: {income:.0f} so'm\n"
        f"   💵 Naqd: {payment_totals['naqd']:.0f} so'm\n"
        f"   💳 Plastik: {payment_totals['plastik']:.0f} so'm\n"
        f"💸 Chiqim: {expense:.0f} so'm\n"
        f"📈 Balans: <b>{balance:.0f} so'm</b>\n\n"
        f"📦 Skladdagi mahsulotlar soni: {len(products)} xil\n"
        f"📦 Sklad qiymati: {total_stock_value:.0f} so'm\n\n"
        f"📒 Umumiy qarzdorlik: {total_debt:.0f} so'm"
    )
    await message.answer(text, parse_mode="HTML")


@router.message(F.text == "🏢 Filial bo'yicha hisobot")
async def branch_report_start(message: Message):
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    branches = await db.get_branches(shop_id)
    if not branches:
        await message.answer(
            "Hozircha filial qo'shilmagan, shuning uchun filial bo'yicha "
            "hisobot mavjud emas. Avval \"🏢 Filiallar\" bo'limidan filial qo'shing."
        )
        return

    await message.answer(
        "Qaysi filial bo'yicha hisobot ko'rmoqchisiz?",
        reply_markup=kb.report_branch_kb(branches, prefix="rep_branch"),
    )


@router.callback_query(F.data.startswith("rep_branch_"))
async def branch_report_show(callback: CallbackQuery):
    shop_id = await get_shop_id(callback.from_user.id)
    if shop_id is None:
        await callback.answer("Bu bo'lim faqat do'kon egalari uchun.", show_alert=True)
        return

    raw = callback.data[len("rep_branch_"):]
    branch_id = int(raw)  # bu yerda "all" bo'lmaydi - kb.report_branch_kb "0" yoki filial id beradi

    branch_name = "🏠 Bosh filial"
    if branch_id:
        branch = await db.get_branch(shop_id, branch_id)
        if not branch:
            await callback.answer("Filial topilmadi.", show_alert=True)
            return
        branch_name = f"🏢 {branch['name']}"

    income, expense = await db.get_totals(shop_id, branch_id=branch_id)
    balance = income - expense
    total_debt = await db.get_total_debt(shop_id, branch_id=branch_id)
    payment_totals = await db.get_payment_method_totals(shop_id, "income", branch_id=branch_id)

    text = (
        f"{branch_name} — <b>hisobot</b>\n\n"
        f"💰 Kirim: {income:.0f} so'm\n"
        f"   💵 Naqd: {payment_totals['naqd']:.0f} so'm\n"
        f"   💳 Plastik: {payment_totals['plastik']:.0f} so'm\n"
        f"💸 Chiqim: {expense:.0f} so'm\n"
        f"📈 Balans: <b>{balance:.0f} so'm</b>\n\n"
        f"📒 Qarzdorlik: {total_debt:.0f} so'm"
    )
    await callback.message.answer(text, parse_mode="HTML")
    await callback.answer()


def _format_branch_comparison(rows: list) -> str:
    """FILIALLAR SOLISHTIRUVI - 11-BOSQICH: db.get_branch_comparison()
    natijasini o'qishga qulay Telegram xabariga aylantiradi."""
    if not rows:
        return (
            "🆚 <b>Filiallar solishtiruvi</b>\n\n"
            "Hozircha solishtirish uchun yetarli ma'lumot yo'q."
        )

    medals = ["🥇", "🥈", "🥉"]
    blocks = [f"🆚 <b>Filiallar solishtiruvi</b> (foyda bo'yicha)"]
    for i, r in enumerate(rows):
        rank = medals[i] if i < len(medals) else f"{i + 1}."
        blocks.append(
            f"{rank} <b>{r['name']}</b>\n"
            f"   🛒 Savdo: {r['sales_count']} ta chek\n"
            f"   💰 Foyda: {r['profit']:.0f} so'm\n"
            f"   💵 Kirim: {r['income']:.0f} so'm  |  💸 Chiqim: {r['expense']:.0f} so'm\n"
            f"   📈 Balans: {r['balance']:.0f} so'm"
        )
    return "\n\n".join(blocks)


@router.message(F.text == "🆚 Filiallar solishtiruvi")
async def branch_comparison_report(message: Message):
    """FILIALLAR SOLISHTIRUVI - 11/12-BOSQICH: "🏢 Filial bo'yicha
    hisobot"dan farqli o'laroq (u yerda faqat BITTA filial tanlab ko'riladi),
    bu yerda BARCHA filiallar (va agar Bosh filialda ham faoliyat bo'lsa -
    u ham) BITTA xabarda, foyda bo'yicha kamayish tartibida solishtiriladi."""
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    branches = await db.get_branches(shop_id)
    if not branches:
        await message.answer(
            "Hozircha filial qo'shilmagan, shuning uchun solishtirish uchun "
            "hech narsa yo'q. Avval \"🏢 Filiallar\" bo'limidan filial qo'shing."
        )
        return

    rows = await db.get_branch_comparison(shop_id)
    await message.answer(_format_branch_comparison(rows), parse_mode="HTML")


def _format_seller_comparison(rows: list) -> str:
    if not rows:
        return (
            "🆚 <b>Sotuvchilar solishtiruvi</b>\n\n"
            "Hozircha solishtirish uchun yetarli ma'lumot yo'q."
        )

    medals = ["🥇", "🥈", "🥉"]
    blocks = ["🆚 <b>Sotuvchilar solishtiruvi</b> (foyda bo'yicha)"]
    for i, r in enumerate(rows):
        rank = medals[i] if i < len(medals) else f"{i + 1}."
        blocks.append(
            f"{rank} <b>{r['name']}</b>\n"
            f"   🛒 Savdo: {r['sales_count']} ta chek\n"
            f"   💰 Foyda: {r['profit']:.0f} so'm\n"
            f"   💵 Kirim: {r['income']:.0f} so'm  |  💸 Chiqim: {r['expense']:.0f} so'm\n"
            f"   📈 Balans: {r['balance']:.0f} so'm"
        )
    return "\n\n".join(blocks)


@router.message(F.text == "🆚 Sotuvchilar solishtiruvi")
async def seller_comparison_report(message: Message):
    """Har bir sotuvchi (va do'kon egasining o'zi) qilgan savdo/foyda/
    kirim-chiqimni bitta xabarda, foyda bo'yicha kamayish tartibida solishtiradi."""
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    rows = await db.get_seller_comparison(shop_id)
    await message.answer(_format_seller_comparison(rows), parse_mode="HTML")


_UZ_MONTHS = {
    "01": "Yanvar", "02": "Fevral", "03": "Mart", "04": "Aprel", "05": "May", "06": "Iyun",
    "07": "Iyul", "08": "Avgust", "09": "Sentabr", "10": "Oktabr", "11": "Noyabr", "12": "Dekabr",
}


def _format_month_uz(month_key: str) -> str:
    """'YYYY-MM' ni 'Iyul 2026' ko'rinishiga o'tkazadi. Format noto'g'ri
    bo'lsa - o'zgarishsiz qaytaradi."""
    try:
        year, month = month_key.split("-")
        return f"{_UZ_MONTHS.get(month, month)} {year}"
    except ValueError:
        return month_key


@router.message(F.text == "📈 Oylik prognoz")
async def monthly_forecast_report(message: Message):
    """OYLIK FOYDA PROGNOZI - 14-BOSQICH: oxirgi 6 oylik savdo/foyda
    tarixini va (agar kamida bitta TO'LIQ tugagan oy bo'lsa) keyingi oy
    uchun oddiy o'rtachaga asoslangan prognozni ko'rsatadi. Batafsil
    o'sish/pasayish tahlili (trend) - 15-bosqichda alohida qo'shiladi,
    bu yerda faqat "oddiy hisoblash" (oxirgi oylar o'rtachasi) bor."""
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    history = await db.get_monthly_profit_history(shop_id, months=6)
    forecast = await db.get_profit_forecast(shop_id)

    blocks = ["📈 <b>Oylik foyda — so'nggi 6 oy</b>"]
    month_lines = [
        f"• {_format_month_uz(h['month'])}: {h['sales_count']} ta chek, foyda {h['profit']:.0f} so'm"
        for h in history
    ]
    blocks.append("\n".join(month_lines))

    any_data = any(h["sales_count"] for h in history)
    if not any_data:
        blocks.append("Hozircha savdo tarixi yo'q.")
    elif forecast:
        blocks.append(
            f"🔮 <b>Prognoz — {_format_month_uz(forecast['forecast_month'])}</b>\n"
            f"Oxirgi {forecast['based_on_months']} oy o'rtachasi asosida taxminiy foyda: "
            f"<b>{forecast['forecast_profit']:.0f} so'm</b>\n"
            f"(o'rtacha savdo summasi: {forecast['avg_sales_total']:.0f} so'm)"
        )
    else:
        blocks.append(
            "🔮 Prognoz uchun kamida bitta TO'LIQ tugagan oylik savdo tarixi kerak "
            "(joriy oy hali tugamagani uchun hisobga olinmaydi)."
        )

    await message.answer("\n\n".join(blocks), parse_mode="HTML")


def _trend_direction(change_percent) -> str:
    """TREND TAHLILI - 15-BOSQICH: o'zgarish foiziga qarab yo'nalish
    belgisini qaytaradi. ±5% ichidagi o'zgarish "deyarli o'zgarishsiz"
    (➡️) deb hisoblanadi - kichik tebranishlarni trend sifatida
    ko'rsatmaslik uchun."""
    if change_percent is None:
        return "▪️"
    if change_percent > 5:
        return "📈"
    if change_percent < -5:
        return "📉"
    return "➡️"


def _compute_trend(series: list) -> list:
    """[{"label": ..., "value": ...}, ...] ketma-ketligidagi har bir davr
    uchun OLDINGI davrga nisbatan o'zgarish foizini hisoblab qo'shadi.
    Birinchi davr uchun solishtiradigan narsa yo'q - change_percent None."""
    result = []
    prev_value = None
    for item in series:
        value = item["value"]
        change_percent = None
        if prev_value is not None and prev_value != 0:
            change_percent = (value - prev_value) / prev_value * 100
        result.append({**item, "change_percent": change_percent})
        prev_value = value
    return result


def _format_trend_series(title: str, trend: list) -> str:
    lines = [title]
    for t in trend:
        arrow = _trend_direction(t["change_percent"])
        change_text = f" ({t['change_percent']:+.0f}%)" if t["change_percent"] is not None else ""
        lines.append(f"{arrow} {t['label']}: {t['value']:.0f} so'm{change_text}")
    return "\n".join(lines)


@router.message(F.text == "📉 Trend tahlili")
async def trend_analysis_report(message: Message):
    """TREND TAHLILI - 15-BOSQICH: foydaning OY va HAFTA kesimida qanday
    o'zgarib borayotganini (o'sish 📈 / pasayish 📉 / deyarli o'zgarishsiz ➡️,
    har birida oldingi davrga nisbatan foiz bilan) ko'rsatadi. 14-bosqichdagi
    "oddiy prognoz"dan farqi - bu yerda kelajak taxmin qilinmaydi, faqat
    O'TGAN davrlar orasidagi dinamika ko'rsatiladi."""
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    monthly = await db.get_monthly_profit_history(shop_id, months=6)
    weekly = await db.get_weekly_profit_history(shop_id, weeks=8)

    if not any(h["sales_count"] for h in monthly):
        await message.answer("Hozircha savdo tarixi yo'q, trend tahlili uchun ma'lumot yetarli emas.")
        return

    monthly_series = [{"label": _format_month_uz(h["month"]), "value": h["profit"]} for h in monthly]
    weekly_series = [
        {
            "label": (
                f"{datetime.strptime(h['week_start'], '%Y-%m-%d').strftime('%d.%m')}–"
                f"{datetime.strptime(h['week_end'], '%Y-%m-%d').strftime('%d.%m')}"
            ),
            "value": h["profit"],
        }
        for h in weekly
    ]

    text = (
        _format_trend_series("📅 <b>Oylik foyda trendi</b> (so'nggi 6 oy)", _compute_trend(monthly_series))
        + "\n\n"
        + _format_trend_series("🗓 <b>Haftalik foyda trendi</b> (so'nggi 8 hafta)", _compute_trend(weekly_series))
    )
    await message.answer(text, parse_mode="HTML")


def _format_top_products(top_selling, top_profit, scope_name: str = ""):
    lines = []
    if scope_name:
        lines.append(f"{scope_name}\n")
    lines.append("🏆 <b>Top 10 - eng ko'p sotilgan</b>\n")
    if top_selling:
        for i, r in enumerate(top_selling, 1):
            lines.append(f"{i}. {r['name']} — {r['total_qty']:.0f} dona ({r['total_sum']:.0f} so'm)")
    else:
        lines.append("Ma'lumot yo'q.")

    lines.append("\n💰 <b>Top 10 - eng ko'p foyda keltirgan</b>\n")
    if top_profit:
        for i, r in enumerate(top_profit, 1):
            lines.append(f"{i}. {r['name']} — {r['total_profit']:.0f} so'm foyda")
    else:
        lines.append("Ma'lumot yo'q.")

    return "\n".join(lines)


async def _send_top_products(target: Message, shop_id: int, branch_id, scope_name: str = ""):
    top_selling = await db.get_top_selling_products(shop_id, limit=10, branch_id=branch_id)
    top_profit = await db.get_top_profit_products(shop_id, limit=10, branch_id=branch_id)
    await target.answer(_format_top_products(top_selling, top_profit, scope_name), parse_mode="HTML")


@router.message(F.text == "🏆 Top mahsulotlar")
async def top_products_report(message: Message):
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    branches = await db.get_branches(shop_id)
    if not branches:
        # Filial umuman yo'q bo'lsa, kesim tanlashning ma'nosi yo'q -
        # to'g'ridan-to'g'ri umumiy natijani ko'rsatamiz.
        await _send_top_products(message, shop_id, branch_id=None)
        return

    await message.answer(
        "Qaysi kesimda top mahsulotlarni ko'rmoqchisiz?",
        reply_markup=kb.report_branch_kb(branches, prefix="rep_top", include_all=True),
    )


@router.callback_query(F.data.startswith("rep_top_"))
async def top_products_show(callback: CallbackQuery):
    shop_id = await get_shop_id(callback.from_user.id)
    if shop_id is None:
        await callback.answer("Bu bo'lim faqat do'kon egalari uchun.", show_alert=True)
        return

    raw = callback.data[len("rep_top_"):]
    if raw == "all":
        branch_id = None
        scope_name = "🌐 Umumiy (barcha filiallar)"
    else:
        branch_id = int(raw)
        if branch_id:
            branch = await db.get_branch(shop_id, branch_id)
            if not branch:
                await callback.answer("Filial topilmadi.", show_alert=True)
                return
            scope_name = f"🏢 {branch['name']}"
        else:
            scope_name = "🏠 Bosh filial"

    await _send_top_products(callback.message, shop_id, branch_id, scope_name)
    await callback.answer()


@router.message(F.text == "🔔 Kunlik hisobot")
async def daily_report_settings(message: Message):
    """KUNLIK HISOBOT - 7-BOSQICH: owner har kuni belgilangan vaqtda
    (config.DAILY_REPORT_HOUR:MINUTE) avtomatik yuboriladigan kunlik
    hisobotni o'zi yoqib/o'chirib qo'ya oladi."""
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    enabled = await db.get_daily_report_enabled(shop_id)
    status = "🔔 <b>yoqilgan</b>" if enabled else "🔕 <b>o'chirilgan</b>"
    await message.answer(
        f"Har kuni soat <b>{config.DAILY_REPORT_HOUR:02d}:{config.DAILY_REPORT_MINUTE:02d}</b> "
        f"da avtomatik yuboriladigan kunlik hisobot hozir {status}.",
        parse_mode="HTML",
        reply_markup=kb.daily_report_settings_kb(enabled),
    )


@router.callback_query(F.data.startswith("daily_report_toggle:"))
async def daily_report_toggle(callback: CallbackQuery):
    shop_id = await get_shop_id(callback.from_user.id)
    if shop_id is None:
        await callback.answer("Bu bo'lim faqat do'kon egalari uchun.", show_alert=True)
        return

    turn_on = callback.data.split(":", 1)[1] == "on"
    await db.set_daily_report_enabled(shop_id, turn_on)

    status = "🔔 <b>yoqildi</b>" if turn_on else "🔕 <b>o'chirildi</b>"
    await callback.message.edit_text(f"Kunlik hisobot {status}.", parse_mode="HTML")
    await callback.answer()


@router.message(F.text == "🚨 Shubhali holatlar")
async def suspicious_alert_settings(message: Message):
    """SHUBHALI HOLATLAR - 10-BOSQICH: owner savdo/kirim-chiqim kiritilganda
    real vaqtda aniqlanadigan shubhali holatlar haqida DARHOL keladigan
    Telegram ogohlantirishini o'zi yoqib/o'chirib qo'ya oladi. O'chirilgan
    taqdirda ham tekshiruv davom etadi - faqat serverdagi logga yoziladi,
    Telegram xabari kelmaydi."""
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    enabled = await db.get_suspicious_alert_enabled(shop_id)
    status = "🚨 <b>yoqilgan</b>" if enabled else "🔕 <b>o'chirilgan</b>"
    await message.answer(
        "Savdo yoki kirim-chiqim kiritilganda (manfiy qoldiq, tannarxdan past "
        "sotish, katta chegirma, g'ayrioddiy miqdor/summasi va h.k.) darhol "
        f"keladigan shubhali holat ogohlantirishi hozir {status}.",
        parse_mode="HTML",
        reply_markup=kb.suspicious_alert_settings_kb(enabled),
    )


@router.callback_query(F.data.startswith("suspicious_alert_toggle:"))
async def suspicious_alert_toggle(callback: CallbackQuery):
    shop_id = await get_shop_id(callback.from_user.id)
    if shop_id is None:
        await callback.answer("Bu bo'lim faqat do'kon egalari uchun.", show_alert=True)
        return

    turn_on = callback.data.split(":", 1)[1] == "on"
    await db.set_suspicious_alert_enabled(shop_id, turn_on)

    status = "🚨 <b>yoqildi</b>" if turn_on else "🔕 <b>o'chirildi</b>"
    await callback.message.edit_text(f"Shubhali holatlar ogohlantiruvi {status}.", parse_mode="HTML")
    await callback.answer()


@router.message(F.text == "🗂 Audit jurnali")
async def audit_journal(message: Message):
    """Kim, qachon nima qilgani (kirim/chiqim, savdo bekor qilish, mahsulot
    o'chirish va h.k.) - eng oxirgi 30 ta yozuv, yangisidan boshlab."""
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    rows = await db.get_audit_log(shop_id, limit=30)
    if not rows:
        await message.answer("Audit jurnalida hali yozuv yo'q.")
        return

    lines = ["🗂 <b>Audit jurnali</b> (oxirgi 30 ta amal)\n"]
    for r in rows:
        date = r["created_at"][:16]
        lines.append(f"🕒 {date} — <b>{r['actor_name']}</b>: {r['action']}")
        if r.get("details"):
            lines.append(f"   {r['details']}")

    await message.answer("\n".join(lines), parse_mode="HTML")


@router.message(F.text == "📥 Excel yuklab olish")
async def export_excel(message: Message):
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    products = await db.get_all_products(shop_id)
    transactions = await db.get_transactions(shop_id, limit=1000)
    debts = await db.get_debts(shop_id, only_unpaid=False)
    branches = await db.get_branches(shop_id)
    branch_names = {b["id"]: b["name"] for b in branches}
    # 9-BOSQICH: Telegramdagi "🗂 Audit jurnali" faqat oxirgi 30 ta yozuvni
    # ko'rsatadi (xabar juda uzun bo'lib ketmasligi uchun) - Excel eksportda
    # esa cheklov yo'q (bu yerda uzunlik muammo emas), shuning uchun ancha
    # ko'proq (oxirgi 2000 ta) yozuv olinadi - shu jumladan barkod
    # o'zgarishlari va Mini App/Excel orqali skladga tovar qo'shilgan
    # holatlar ham shu yerda to'liq ko'rinadi.
    audit_rows = await db.get_audit_log(shop_id, limit=2000)

    def branch_label(branch_id):
        if not branch_id:
            return "Bosh filial"
        return branch_names.get(branch_id, "Bosh filial")

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # ---- Sklad ----
    ws1 = wb.active
    ws1.title = "Sklad"
    ws1.append(["ID", "Nomi", "Narxi", "Miqdori", "Jami qiymat", "Barkod", "Qo'shilgan sana"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
    for p in products:
        ws1.append([
            p["id"], p["name"], p["price"], p["quantity"],
            p["price"] * p["quantity"], p.get("barcode") or "", p["created_at"][:10]
        ])

    # ---- Kirim/Chiqim ----
    ws2 = wb.create_sheet("Kirim-Chiqim")
    ws2.append(["ID", "Turi", "Summasi", "To'lov turi", "Izoh", "Filial", "Sana"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for t in transactions:
        label = "Kirim" if t["type"] == "income" else "Chiqim"
        method_map = {"naqd": "Naqd", "plastik": "Plastik"}
        method_label = method_map.get(t.get("payment_method"), "")
        ws2.append([
            t["id"], label, t["amount"], method_label, t["description"],
            branch_label(t.get("branch_id")), t["created_at"][:16]
        ])

    # ---- Qarz daftar ----
    ws3 = wb.create_sheet("Qarz daftar")
    ws3.append([
        "ID", "Mijoz", "Telefon", "Qarz summasi", "To'langan", "Qolgan",
        "Izoh", "Holati", "Filial", "Olingan sanasi", "Qaytarish sanasi", "Kiritilgan sana"
    ])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for d in debts:
        status = "To'landi" if d["is_paid"] else "Qarzda"
        paid_amount = d.get("paid_amount") or 0
        remaining = d["amount"] - paid_amount
        taken_date = d.get("taken_date") or d["created_at"][:10]
        due_date = d.get("due_date") or ""
        ws3.append([
            d["id"], d["customer_name"], d["phone"], d["amount"], paid_amount, remaining,
            d["description"], status, branch_label(d.get("branch_id")),
            taken_date, due_date, d["created_at"][:10]
        ])

    # ---- Audit jurnali ----
    ws4 = wb.create_sheet("Audit jurnali")
    ws4.append(["ID", "Sana", "Kim", "Amal", "Tafsilot"])
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in audit_rows:
        ws4.append([
            r["id"], r["created_at"][:16], r["actor_name"], r["action"], r.get("details") or "",
        ])

    # Ustun kengligini avtomoslashtirish
    for ws in (ws1, ws2, ws3, ws4):
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 3

    filename = f"hisobot_{config.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)

    await message.answer_document(FSInputFile(filepath, filename=filename))
    os.remove(filepath)


# ---------- DB FAYLNI ZAXIRALASH / TIKLASH ----------
# Railway'da Volume ulanmagan bo'lsa, redeploy vaqtida baza tozalanadi va
# barcha do'konlarning ma'lumotlarini qaytadan kiritish kerak bo'ladi. Bu -
# BUTUN bazani (barcha do'konlar birdaniga) qamrab oladigan texnik zaxira
# vositasi, shuning uchun endi faqat BOSH ADMINGA ochiq - alohida do'kon
# egasi bunga kira olmaydi (aks holda u boshqa do'konlarning ma'lumotlarini
# ko'rib/almashtirib qo'yishi mumkin edi).

@router.message(F.text == "🗄 DB faylni yuklab olish")
async def export_db(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu bo'lim faqat bosh admin uchun.")
        return

    if not os.path.exists(config.DB_PATH):
        await message.answer("❌ Baza fayli topilmadi.")
        return

    filename = os.path.basename(config.DB_PATH) or "shop.db"
    await message.answer_document(
        FSInputFile(config.DB_PATH, filename=filename),
        caption=(
            "🗄 Joriy baza fayli (BARCHA do'konlar).\n"
            "Buni saqlab qo'ying — baza tozalanib qolsa (masalan redeploy'da), "
            "\"📤 DB faylni tiklash\" tugmasi orqali shu faylni qayta yuklab, "
            "barcha do'konlarni qaytadan kiritmasdan tiklashingiz mumkin."
        ),
    )


@router.message(F.text == "📤 DB faylni tiklash")
async def restore_db_start(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Bu bo'lim faqat bosh admin uchun.")
        return

    await state.set_state(RestoreDB.waiting_file)
    await message.answer(
        "⚠️ <b>Diqqat!</b> Yuboradigan faylingiz joriy bazadagi BARCHA do'konlarning "
        "ma'lumotlarini (mahsulotlar, kirim-chiqim, qarzlar, do'kon egalari) to'liq almashtiradi.\n\n"
        "Avval joriy bazani zaxiralab olmoqchi bo'lsangiz, \"🗄 DB faylni yuklab olish\" "
        "tugmasini bosing.\n\n"
        "Tiklash uchun oldin saqlab qo'ygan .db faylni hujjat sifatida yuboring "
        "(bekor qilish uchun \"⬅️ Orqaga\" tugmasini bosing):",
        parse_mode="HTML",
    )


@router.message(RestoreDB.waiting_file, F.document)
async def restore_db_file(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return

    document = message.document
    tmp_path = f"/tmp/restore_{document.file_unique_id}.db"
    await message.bot.download(document.file_id, destination=tmp_path)

    valid = False
    try:
        async with aiosqlite.connect(tmp_path) as test_db:
            cursor = await test_db.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='products'"
            )
            row = await cursor.fetchone()
            valid = row is not None
    except Exception:
        valid = False

    if not valid:
        os.remove(tmp_path)
        await message.answer(
            "❌ Bu fayl to'g'ri baza fayli emasga o'xshaydi. Qaytadan urinib ko'ring "
            "yoki \"⬅️ Orqaga\" tugmasini bosing."
        )
        return

    db_dir = os.path.dirname(config.DB_PATH)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    shutil.move(tmp_path, config.DB_PATH)

    # Tiklangan fayl eski struktura bilan bo'lishi mumkin (masalan
    # subscription_status kabi keyinroq qo'shilgan ustunlar yo'q bo'lishi
    # mumkin) - shuning uchun joylashtirilgach migratsiyalarni qayta
    # ishga tushiramiz, aks holda keyingi so'rovlar "no column named ..."
    # xatosi bilan yiqiladi.
    try:
        await db.init_db()
    except Exception:
        await message.answer(
            "⚠️ Baza fayli ko'chirildi, lekin migratsiyani qayta ishga tushirishda xato "
            "yuz berdi. Botni qo'lda qayta ishga tushiring (restart)."
        )
        await state.clear()
        return

    await state.clear()
    await message.answer(
        "✅ Baza muvaffaqiyatli tiklandi (barcha do'konlar). Qaytadan kiritishning hojati yo'q.",
        reply_markup=kb.main_menu("admin"),
    )


@router.message(RestoreDB.waiting_file)
async def restore_db_wrong_type(message: Message):
    await message.answer(
        "Iltimos, .db faylni hujjat (document) sifatida yuboring, "
        "yoki bekor qilish uchun \"⬅️ Orqaga\" tugmasini bosing."
    )
