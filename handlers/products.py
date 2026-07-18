import logging
import os

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

import config
import database as db
import keyboards as kb
import alerts
from access_control import get_shop_id, is_owner_level, get_role

router = Router()


class AddProduct(StatesGroup):
    name = State()
    category = State()
    new_category = State()
    price = State()
    sell_price = State()
    min_price = State()
    quantity = State()
    alert_quantity = State()
    photo = State()


class CategoryManage(StatesGroup):
    new_name = State()


class SetDiscount(StatesGroup):
    """Mahsulotga vaqtinchalik chegirma narx belgilash - FAQAT do'kon
    egasiga ruxsat etilgan (har bir handlerda is_owner_level tekshiriladi)."""
    price = State()
    days = State()


class ExcelFill(StatesGroup):
    """Skladni Excel fayl orqali to'ldirish - shablon yuborilgach shu
    holatga o'tiladi va navbatdagi hujjat (.xlsx) shu shop_id'ga qarab
    qayta ishlanadi. target_shop_id state ma'lumotida saqlanadi - odatda
    foydalanuvchining o'z do'koni, lekin admin boshqa eganing do'konini
    to'ldirayotganda shu yerga o'sha eganing shop_id'si yoziladi
    (handlers/users.py'dagi admin_sklad_excel_start_cb orqali)."""
    waiting_file = State()


class SearchProduct(StatesGroup):
    query = State()


class AddRestockItem(StatesGroup):
    name = State()
    note = State()


class RestockPurchase(StatesGroup):
    quantity = State()
    price = State()
    sell_price = State()
    min_price = State()
    alert_quantity = State()


# Bu bo'lim tugmalari faqat do'kon egalariga ko'rsatiladi (bosh adminning o'z
# do'koni yo'q). Shunga qaramay, har bir handler shop_id'ni qayta tekshiradi -
# bosh admin adashib shu bo'limga kirib qolsa ham, hech qanday do'kon
# ma'lumotiga ega bo'lmaydi.

async def _require_shop(message: Message):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await message.answer("Bu bo'lim faqat do'kon egalari uchun.")
    return shop_id


async def _require_shop_cb(callback: CallbackQuery):
    shop_id = await get_shop_id(callback.from_user.id)
    if shop_id is None:
        await callback.answer("Bu bo'lim faqat do'kon egalari uchun.", show_alert=True)
    return shop_id


# ---------- SKLADNI EXCEL BILAN TO'LDIRISH ----------
# Do'kon egasi (yoki admin - handlers/users.py orqali, boshqa eganing
# do'koni uchun) skladni bittalab emas, bitta Excel fayl orqali to'ldira
# oladi: avval bot shablon (namuna) faylni o'zi tashlab beradi, keyin
# to'ldirilgan fayl hujjat sifatida qaytarib yuborilsa, undagi har bir
# qator mahsulot sifatida qo'shiladi/yangilanadi. send_products_excel_template
# va process_products_excel_file funksiyalari qasddan alohida ajratilgan
# (router handlerlariga bog'lanmagan) - handlers/users.py ularni import
# qilib, admin uchun ham xuddi shu mantiqni ishlatadi.

EXCEL_TEMPLATE_HEADERS = [
    "Nomi", "Tannarx", "Sotuv narxi", "Eng past narx", "Miqdori",
    "Ogohlantirish chegarasi", "Bo'lim",
]


def _build_products_template_file() -> str:
    """Bo'sh (namunali bitta qator bilan) Excel shablonini /tmp'ga
    yozadi va fayl yo'lini qaytaradi. Faqat \"Nomi\", \"Tannarx\" va
    \"Miqdori\" ustunlari majburiy - qolganlari bo'sh qoldirilishi mumkin."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sklad"
    ws.append(EXCEL_TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.append(["Coca-Cola 1.5L", 8000, 10000, 9000, 50, 5, "Ichimliklar"])
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 3

    filepath = f"/tmp/sklad_shablon_{os.getpid()}_{id(wb)}.xlsx"
    wb.save(filepath)
    return filepath


async def send_products_excel_template(message: Message):
    """Shablon faylni yuboradi. Chaqiruvchi (owner yoki admin) buni
    to'ldirib, keyin ExcelFill.waiting_file holatida qaytarib yuboradi."""
    filepath = _build_products_template_file()
    try:
        await message.answer_document(
            FSInputFile(filepath, filename="sklad_shablon.xlsx"),
            caption=(
                "📥 <b>Sklad shablon fayli</b>\n\n"
                "Ustunlar: <b>Nomi</b>, <b>Tannarx</b> va <b>Miqdori</b> - majburiy. "
                "Sotuv narxi, Eng past narx, Ogohlantirish chegarasi va Bo'lim - ixtiyoriy "
                "(bo'lim nomi hali mavjud bo'lmasa, avtomatik yaratiladi).\n\n"
                "Mavjud nomdagi mahsulot qatorga tushsa - miqdori qo'shiladi va tannarxi "
                "o'rtacha (weighted average) qayta hisoblanadi, yangi nom bo'lsa - yangi "
                "mahsulot sifatida qo'shiladi.\n\n"
                "To'ldirib bo'lgach, shu faylni <b>hujjat</b> sifatida shu yerga qaytarib yuboring."
            ),
            parse_mode="HTML",
        )
    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass


async def process_products_excel_file(shop_id: int, file_path: str) -> dict:
    """Yuborilgan Excel faylini o'qib, shop_id do'konining skladiga
    qo'llaydi. Qaytaradi: {"added": int, "updated": int, "errors": [str, ...]}.

    Nomi bo'yicha (katta-kichik harfsiz) mos kelgan mahsulot bo'lsa -
    database.update_product_purchase() orqali miqdor qo'shiladi va tannarx
    o'rtacha (weighted average) qilib qayta hisoblanadi; sotuv/eng past
    narx/ogohlantirish ustunlari faylda bo'sh qoldirilgan bo'lsa, mahsulotning
    eski qiymati saqlab qolinadi (update_product_purchase berilgan qiymatni
    so'zsiz yozib qo'yadi, shuning uchun bu yerda avval fallback qilinadi)."""
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception:
        return {"added": 0, "updated": 0, "errors": ["Fayl xato yoki buzilgan - .xlsx formatida ekanini tekshiring."]}

    ws = wb.active
    categories = await db.get_categories(shop_id)
    cat_by_name = {c["name"].strip().lower(): c["id"] for c in categories}
    existing_products = await db.get_all_products(shop_id)
    existing_by_name = {p["name"].strip().lower(): p for p in existing_products}

    added = 0
    updated = 0
    errors = []

    def _num(row, idx):
        if idx >= len(row) or row[idx] in (None, ""):
            return None, True
        try:
            return float(row[idx]), True
        except (ValueError, TypeError):
            return None, False

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        name = str(row[0]).strip() if row[0] is not None else ""
        if not name:
            continue

        price, price_ok = _num(row, 1)
        if not price_ok:
            errors.append(f"{row_idx}-qator ({name}): Tannarx raqam emas")
            continue
        quantity, qty_ok = _num(row, 4)
        if not qty_ok:
            errors.append(f"{row_idx}-qator ({name}): Miqdori raqam emas")
            continue
        if price is None or quantity is None:
            errors.append(f"{row_idx}-qator ({name}): Tannarx va Miqdori majburiy")
            continue

        sell_price, sell_ok = _num(row, 2)
        if not sell_ok:
            errors.append(f"{row_idx}-qator ({name}): Sotuv narxi raqam emas")
            continue
        min_price, min_ok = _num(row, 3)
        if not min_ok:
            errors.append(f"{row_idx}-qator ({name}): Eng past narx raqam emas")
            continue
        alert_quantity, alert_ok = _num(row, 5)
        if not alert_ok:
            errors.append(f"{row_idx}-qator ({name}): Ogohlantirish chegarasi raqam emas")
            continue

        category_id = None
        if len(row) > 6 and row[6] not in (None, ""):
            cat_name = str(row[6]).strip()
            key = cat_name.lower()
            category_id = cat_by_name.get(key)
            if category_id is None:
                category = await db.add_category(shop_id, cat_name)
                category_id = category["id"]
                cat_by_name[key] = category_id

        existing = existing_by_name.get(name.lower())
        if existing:
            final_sell_price = sell_price if sell_price is not None else existing.get("sell_price")
            final_min_price = min_price if min_price is not None else existing.get("min_price")
            final_alert_qty = alert_quantity if alert_quantity is not None else existing.get("alert_quantity")
            await db.update_product_purchase(
                shop_id, existing["id"], quantity, price,
                final_sell_price, final_min_price, final_alert_qty,
            )
            if category_id is not None:
                await db.set_product_category(shop_id, existing["id"], category_id)
            existing["quantity"] = existing["quantity"] + quantity
            updated += 1
        else:
            await db.add_product(
                shop_id, name, price, quantity, None,
                sell_price=sell_price, min_price=min_price,
                alert_quantity=alert_quantity, category_id=category_id,
            )
            existing_by_name[name.lower()] = {
                "id": None, "quantity": quantity, "price": price,
                "sell_price": sell_price, "min_price": min_price,
                "alert_quantity": alert_quantity,
            }
            added += 1

    return {"added": added, "updated": updated, "errors": errors}


def excel_result_text(result: dict, shop_label: str = "") -> str:
    """process_products_excel_file() natijasini foydalanuvchiga
    ko'rsatiladigan matnga aylantiradi. shop_label - admin oqimida
    \"do'kon: <ism>\" kabi qo'shimcha izoh uchun (owner oqimida bo'sh)."""
    text = "✅ <b>Sklad Excel fayl orqali to'ldirildi</b>"
    if shop_label:
        text += f" ({shop_label})"
    text += f"\n\n➕ Yangi qo'shildi: {result['added']} ta\n🔄 Yangilandi: {result['updated']} ta"
    if result["errors"]:
        shown = result["errors"][:10]
        text += f"\n\n⚠️ Quyidagi qatorlarda xatolik bor, ular o'tkazib yuborildi:\n" + "\n".join(shown)
        if len(result["errors"]) > len(shown):
            text += f"\n... yana {len(result['errors']) - len(shown)} ta xatolik."
    return text


@router.message(F.text == "📊 Excel bilan to'ldirish")
async def excel_fill_start(message: Message, state: FSMContext):
    shop_id = await _require_shop(message)
    if shop_id is None:
        return
    await state.update_data(target_shop_id=shop_id)
    await state.set_state(ExcelFill.waiting_file)
    await send_products_excel_template(message)


@router.message(ExcelFill.waiting_file, F.document)
async def excel_fill_file(message: Message, state: FSMContext):
    data = await state.get_data()
    shop_id = data.get("target_shop_id")
    if shop_id is None:
        await state.clear()
        return

    document = message.document
    if not (document.file_name or "").lower().endswith((".xlsx", ".xlsm")):
        await message.answer("Iltimos, .xlsx formatidagi Excel faylni yuboring.")
        return

    tmp_path = f"/tmp/sklad_fill_{document.file_unique_id}.xlsx"
    await message.bot.download(document.file_id, destination=tmp_path)
    try:
        result = await process_products_excel_file(shop_id, tmp_path)
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    await state.clear()
    role = await get_role(message.from_user.id)
    menu = kb.main_menu("admin") if role == "admin" else kb.sklad_menu()
    await message.answer(excel_result_text(result), parse_mode="HTML", reply_markup=menu)


@router.message(ExcelFill.waiting_file)
async def excel_fill_wrong_type(message: Message):
    await message.answer(
        "Iltimos, to'ldirilgan Excel faylni hujjat (document) sifatida yuboring, "
        "yoki bekor qilish uchun \"⬅️ Orqaga\" tugmasini bosing."
    )


# ---------- MAHSULOT QO'SHISH ----------

@router.message(F.text == "➕ Mahsulot qo'shish")
async def add_product_start(message: Message, state: FSMContext):
    if await _require_shop(message) is None:
        return
    await state.set_state(AddProduct.name)
    await message.answer("Mahsulot nomini kiriting:")


@router.message(AddProduct.name)
async def add_product_name(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await state.clear()
        return

    name = message.text.strip()
    existing = await db.find_product_by_name(shop_id, name)
    if existing:
        # Bunday nomli mahsulot (katta-kichik harfidan qat'i nazar) skladda
        # allaqachon bor - yangi qator ochmaymiz, balki shu mahsulotga miqdor
        # qo'shish (restock) oqimiga o'tkazamiz.
        await state.set_state(RestockPurchase.quantity)
        await state.update_data(
            restock_type="product", product_id=existing["id"], name=existing["name"]
        )
        await message.answer(
            f"ℹ️ Bunday mahsulot skladda mavjud: <b>{existing['name']}</b> "
            f"(hozir {existing['quantity']:.0f} dona bor).\n"
            f"Kiritayotgan miqdoringiz shu mahsulotga qo'shiladi.\n\n"
            f"Necha dona sotib olindi?",
            parse_mode="HTML",
        )
        return

    await state.update_data(name=name)
    categories = await db.get_categories(shop_id)
    await state.set_state(AddProduct.category)
    if categories:
        await message.answer(
            "Qaysi bo'limga tegishli? Mavjudlaridan tanlang yoki yangisini qo'shing:",
            reply_markup=kb.category_pick_kb(categories),
        )
    else:
        await message.answer(
            "Hozircha bo'lim yo'q. Yangi bo'lim qo'shing yoki bo'limsiz davom eting:",
            reply_markup=kb.category_pick_kb(categories),
        )


async def _ask_price(message: Message, state: FSMContext):
    await state.set_state(AddProduct.price)
    await message.answer("Tannarxini kiriting (necha so'mga tushdi, faqat raqam):")


@router.callback_query(AddProduct.category, F.data == "cat_pick_new")
async def add_product_category_new(callback: CallbackQuery, state: FSMContext):
    if await _require_shop_cb(callback) is None:
        return
    await state.set_state(AddProduct.new_category)
    await callback.message.answer("Yangi bo'lim nomini kiriting:")
    await callback.answer()


@router.message(AddProduct.new_category)
async def add_product_category_new_name(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await state.clear()
        return
    name = message.text.strip()
    if not name:
        await message.answer("Bo'lim nomi bo'sh bo'lishi mumkin emas. Qaytadan kiriting:")
        return
    category = await db.add_category(shop_id, name)
    await state.update_data(category_id=category["id"])
    await message.answer(f"✅ \"{category['name']}\" bo'limi tanlandi.")
    await _ask_price(message, state)


@router.callback_query(AddProduct.category, F.data == "cat_pick_none")
async def add_product_category_none(callback: CallbackQuery, state: FSMContext):
    if await _require_shop_cb(callback) is None:
        return
    await state.update_data(category_id=None)
    await callback.answer()
    await _ask_price(callback.message, state)


@router.callback_query(AddProduct.category, F.data.startswith("cat_pick_"))
async def add_product_category_pick(callback: CallbackQuery, state: FSMContext):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    category_id = int(callback.data.split("_")[-1])
    category = await db.get_category(shop_id, category_id)
    if not category:
        await callback.answer("Bo'lim topilmadi", show_alert=True)
        return
    await state.update_data(category_id=category_id)
    await callback.answer(f"\"{category['name']}\" tanlandi")
    await _ask_price(callback.message, state)


@router.message(AddProduct.price)
async def add_product_price(message: Message, state: FSMContext):
    try:
        price = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 25000")
        return
    await state.update_data(price=price)
    await state.set_state(AddProduct.sell_price)
    await message.answer("Savdo narxini kiriting (necha so'mga sotasiz):")


@router.message(AddProduct.sell_price)
async def add_product_sell_price(message: Message, state: FSMContext):
    try:
        sell_price = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 30000")
        return
    data = await state.get_data()
    if sell_price < data["price"]:
        await message.answer(
            f"❌ Savdo narxi tannarxdan ({data['price']:.0f} so'm) past bo'lishi mumkin emas, "
            f"aks holda zararga ketasiz. Qaytadan kiriting:"
        )
        return
    await state.update_data(sell_price=sell_price)
    await state.set_state(AddProduct.min_price)
    await message.answer("Eng past narxini kiriting (qanchagacha tushirib berish mumkin):")


@router.message(AddProduct.min_price)
async def add_product_min_price(message: Message, state: FSMContext):
    try:
        min_price = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 27000")
        return
    data = await state.get_data()
    if min_price > data["sell_price"]:
        await message.answer(
            f"Eng past narx savdo narxidan ({data['sell_price']:.0f} so'm) katta bo'lmasligi kerak. Qaytadan kiriting:"
        )
        return
    if min_price < data["price"]:
        await message.answer(
            f"❌ Eng past narx tannarxdan ({data['price']:.0f} so'm) past bo'lishi mumkin emas, "
            f"aks holda chegirmada zararga ketasiz. Qaytadan kiriting:"
        )
        return
    await state.update_data(min_price=min_price)
    await state.set_state(AddProduct.quantity)
    await message.answer("Miqdorini kiriting (masalan: 10):")


@router.message(AddProduct.quantity)
async def add_product_quantity(message: Message, state: FSMContext):
    try:
        quantity = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 10")
        return
    await state.update_data(quantity=quantity)
    await state.set_state(AddProduct.alert_quantity)
    await message.answer(
        "Mahsulot nechta qolganda ogohlantirish yuborilsin?\n"
        "(Kerak bo'lmasa 0 deb yozing)"
    )


@router.message(AddProduct.alert_quantity)
async def add_product_alert_quantity(message: Message, state: FSMContext):
    try:
        alert_quantity = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 5 (kerak bo'lmasa 0)")
        return
    await state.update_data(alert_quantity=alert_quantity if alert_quantity > 0 else None)
    await state.set_state(AddProduct.photo)
    await message.answer(
        "Mahsulot rasmini yuboring (yoki rasmsiz davom eting):",
        reply_markup=kb.skip_photo_kb()
    )


@router.message(AddProduct.photo, F.photo)
async def add_product_photo(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await state.clear()
        return
    data = await state.get_data()
    photo_file_id = message.photo[-1].file_id
    channel_message_id = None

    # Rasmni kanalga jo'natib, o'sha yerdagi file_id va message_id'ni olamiz.
    # Shunda rasm botning o'z serverida emas, Telegram kanalida saqlanadi
    # va foydalanuvchi bot bilan chatni o'chirsa ham rasm yo'qolmaydi.
    # message_id'ni saqlab qo'yamiz - mahsulot o'chirilsa yoki miqdori
    # o'zgarsa, kanaldagi postni ham shu orqali yangilaymiz.
    if config.CHANNEL_ID:
        try:
            sent = await message.bot.send_photo(
                chat_id=config.CHANNEL_ID,
                photo=photo_file_id,
                caption=(
                    f"🆕 {data['name']} | Savdo narxi: {data['sell_price']:.0f} so'm | "
                    f"{data['quantity']:.0f} dona"
                ),
            )
            photo_file_id = sent.photo[-1].file_id
            channel_message_id = sent.message_id
        except Exception as e:
            logging.warning(f"Rasmni kanalga yuborib bo'lmadi: {e}")

    await db.add_product(
        shop_id, data["name"], data["price"], data["quantity"], photo_file_id, channel_message_id,
        sell_price=data["sell_price"], min_price=data["min_price"],
        alert_quantity=data.get("alert_quantity"), category_id=data.get("category_id"),
    )
    await state.clear()
    alert_line = ""
    if data.get("alert_quantity"):
        alert_line = f"Ogohlantirish chegarasi: {data['alert_quantity']:.0f} dona\n"
    await message.answer(
        f"✅ Mahsulot qo'shildi:\n<b>{data['name']}</b>\nTannarx: {data['price']:.0f} so'm\n"
        f"Savdo narxi: {data['sell_price']:.0f} so'm\nEng past narx: {data['min_price']:.0f} so'm\n"
        f"Miqdori: {data['quantity']:.0f}\n{alert_line}",
        reply_markup=kb.sklad_menu(),
        parse_mode="HTML"
    )


@router.callback_query(AddProduct.photo, F.data == "skip_photo")
async def add_product_skip_photo(callback: CallbackQuery, state: FSMContext):
    shop_id = await get_shop_id(callback.from_user.id)
    if shop_id is None:
        await state.clear()
        await callback.answer()
        return
    data = await state.get_data()
    await db.add_product(
        shop_id, data["name"], data["price"], data["quantity"], None,
        sell_price=data["sell_price"], min_price=data["min_price"],
        alert_quantity=data.get("alert_quantity"), category_id=data.get("category_id"),
    )
    await state.clear()
    alert_line = ""
    if data.get("alert_quantity"):
        alert_line = f"Ogohlantirish chegarasi: {data['alert_quantity']:.0f} dona\n"
    await callback.message.answer(
        f"✅ Mahsulot qo'shildi:\n<b>{data['name']}</b>\nTannarx: {data['price']:.0f} so'm\n"
        f"Savdo narxi: {data['sell_price']:.0f} so'm\nEng past narx: {data['min_price']:.0f} so'm\n"
        f"Miqdori: {data['quantity']:.0f}\n{alert_line}",
        reply_markup=kb.sklad_menu(),
        parse_mode="HTML"
    )
    await callback.answer()


# ---------- MAHSULOTLAR RO'YXATI (BO'LIM BO'YICHA) ----------

def _product_caption(p: dict) -> str:
    caption = (
        f"<b>{p['name']}</b>\n"
        f"Tannarx: {p['price']:.0f} so'm\n"
    )
    if p.get("sell_price"):
        caption += f"Savdo narxi: {p['sell_price']:.0f} so'm\n"
    if p.get("min_price"):
        caption += f"Eng past narx: {p['min_price']:.0f} so'm\n"
    discount = db.product_discount_info(p)
    if discount:
        caption += (
            f"🏷 Chegirma narxi: {discount['price']:.0f} so'm "
            f"({discount['days_left']} kun qoldi)\n"
        )
    caption += f"Miqdori: {p['quantity']:.0f}"
    if p.get("alert_quantity"):
        caption += f"\nOgohlantirish chegarasi: {p['alert_quantity']:.0f} dona"
    return caption


async def _send_products(message: Message, products):
    """Berilgan mahsulotlar ro'yxatini bittalab, rasm bilan/rasmsiz chiqaradi."""
    for p in products:
        caption = _product_caption(p)
        markup = kb.product_action_kb(
            p["id"],
            category_id=p.get("category_id"),
            has_discount=bool(db.product_discount_info(p)),
        )
        if p["photo_file_id"]:
            await message.answer_photo(
                photo=p["photo_file_id"],
                caption=caption,
                reply_markup=markup,
                parse_mode="HTML"
            )
        else:
            await message.answer(
                caption,
                reply_markup=markup,
                parse_mode="HTML"
            )


@router.message(F.text == "📋 Mahsulotlar ro'yxati")
async def list_products(message: Message):
    shop_id = await _require_shop(message)
    if shop_id is None:
        return
    products = await db.get_all_products(shop_id)
    if not products:
        await message.answer("Sklad hozircha bo'sh.")
        return

    categories = await db.get_categories(shop_id)
    uncategorized_count = await db.get_uncategorized_count(shop_id)
    await message.answer(
        f"Jami {len(products)} ta mahsulot. Qaysi bo'limni ko'rmoqchisiz?",
        reply_markup=kb.category_browse_kb(categories, uncategorized_count),
    )


@router.callback_query(F.data == "cat_view_all")
async def cat_view_all_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    products = await db.get_all_products(shop_id)
    await callback.answer()
    if not products:
        await callback.message.answer("Sklad hozircha bo'sh.")
        return
    await _send_products(callback.message, products)


@router.callback_query(F.data == "cat_view_none")
async def cat_view_none_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    products = await db.get_products_by_category(shop_id, None)
    await callback.answer()
    if not products:
        await callback.message.answer("Bo'limsiz mahsulot yo'q.")
        return
    await _send_products(callback.message, products)


@router.callback_query(F.data.startswith("cat_view_"))
async def cat_view_one_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    category_id = int(callback.data.split("_")[-1])
    category = await db.get_category(shop_id, category_id)
    if not category:
        await callback.answer("Bo'lim topilmadi", show_alert=True)
        return
    products = await db.get_products_by_category(shop_id, category_id)
    await callback.answer()
    if not products:
        await callback.message.answer(f"\"{category['name']}\" bo'limida mahsulot yo'q.")
        return
    await _send_products(callback.message, products)


# ---------- MAHSULOT QIDIRISH ----------

@router.message(F.text == "🔍 Qidirish")
async def search_product_start(message: Message, state: FSMContext):
    if await _require_shop(message) is None:
        return
    await state.set_state(SearchProduct.query)
    await message.answer("Qidirilayotgan mahsulot nomini kiriting:")


@router.message(SearchProduct.query)
async def search_product_query(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await state.clear()
        return
    await state.clear()
    query = message.text.strip()
    results = await db.search_products(shop_id, query)
    if not results:
        await message.answer(f"\"{query}\" bo'yicha hech narsa topilmadi.")
        return
    await message.answer(f"\"{query}\" bo'yicha {len(results)} ta natija topildi:")
    await _send_products(message, results)


# ---------- BO'LIMLARNI BOSHQARISH ----------

@router.message(F.text == "🗂 Bo'limlar")
async def categories_menu(message: Message):
    if await _require_shop(message) is None:
        return
    shop_id = await get_shop_id(message.from_user.id)
    categories = await db.get_categories(shop_id)
    if not categories:
        await message.answer(
            "Hozircha bo'lim yo'q. Yangi bo'lim qo'shishingiz mumkin:",
            reply_markup=kb.category_manage_kb(categories),
        )
        return
    await message.answer(
        "Mavjud bo'limlar (qavsda mahsulotlar soni):",
        reply_markup=kb.category_manage_kb(categories),
    )


@router.callback_query(F.data == "cat_manage_new")
async def category_manage_new_cb(callback: CallbackQuery, state: FSMContext):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    await state.set_state(CategoryManage.new_name)
    await callback.message.answer("Yangi bo'lim nomini kiriting:")
    await callback.answer()


@router.message(CategoryManage.new_name)
async def category_manage_new_name(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await state.clear()
        return
    name = message.text.strip()
    if not name:
        await message.answer("Bo'lim nomi bo'sh bo'lishi mumkin emas. Qaytadan kiriting:")
        return
    await state.clear()
    category = await db.add_category(shop_id, name)
    categories = await db.get_categories(shop_id)
    await message.answer(
        f"✅ \"{category['name']}\" bo'limi tayyor.",
        reply_markup=kb.category_manage_kb(categories),
    )


@router.callback_query(F.data.startswith("cat_noop_"))
async def category_noop_cb(callback: CallbackQuery):
    await callback.answer()


@router.callback_query(F.data.startswith("cat_delete_"))
async def category_delete_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    category_id = int(callback.data.split("_")[-1])
    deleted = await db.delete_category(shop_id, category_id)
    if not deleted:
        await callback.answer("Bo'lim topilmadi", show_alert=True)
        return
    await callback.answer("Bo'lim o'chirildi. Ichidagi mahsulotlar bo'limsiz bo'lib qoldi.", show_alert=True)
    categories = await db.get_categories(shop_id)
    try:
        await callback.message.edit_text(
            "Mavjud bo'limlar (qavsda mahsulotlar soni):" if categories else
            "Hozircha bo'lim yo'q. Yangi bo'lim qo'shishingiz mumkin:",
            reply_markup=kb.category_manage_kb(categories),
        )
    except Exception:
        pass


@router.callback_query(F.data.startswith("del_product_"))
async def delete_product_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    product_id = int(callback.data.split("_")[-1])
    product = await db.get_product(shop_id, product_id)

    await db.delete_product(shop_id, product_id)

    # Kanaldagi rasm postini ham o'chiramiz.
    if product and product.get("channel_message_id") and config.CHANNEL_ID:
        try:
            await callback.bot.delete_message(
                chat_id=config.CHANNEL_ID,
                message_id=product["channel_message_id"],
            )
        except Exception as e:
            logging.warning(f"Kanaldagi postni o'chirib bo'lmadi: {e}")

    await callback.answer("Mahsulot o'chirildi ✅")
    try:
        await callback.message.delete()
    except Exception:
        pass


# ---------- CHEGIRMA (FAQAT DO'KON EGASI) ----------
# Mahsulotga vaqtinchalik chegirma narx belgilash. Muddati (kun soni)
# tugagach database.product_discount_info() uni avtomatik "tugagan" deb
# hisoblay boshlaydi - alohida fon vazifasi/cron shart emas.

@router.callback_query(F.data.startswith("prod_discount_cancel_"))
async def product_discount_cancel_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    if not await is_owner_level(callback.from_user.id):
        await callback.answer("⛔ Bu amal faqat do'kon egasiga ruxsat etilgan.", show_alert=True)
        return
    product_id = int(callback.data.split("_")[-1])
    cleared = await db.clear_product_discount(shop_id, product_id)
    if not cleared:
        await callback.answer("Mahsulot topilmadi", show_alert=True)
        return
    await callback.answer("✅ Chegirma bekor qilindi", show_alert=True)
    try:
        await callback.message.delete()
    except Exception:
        pass


@router.callback_query(F.data.startswith("prod_discount_") & ~F.data.startswith("prod_discount_cancel_"))
async def product_discount_start_cb(callback: CallbackQuery, state: FSMContext):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    if not await is_owner_level(callback.from_user.id):
        await callback.answer("⛔ Bu amal faqat do'kon egasiga ruxsat etilgan.", show_alert=True)
        return
    product_id = int(callback.data.split("_")[-1])
    product = await db.get_product(shop_id, product_id)
    if not product:
        await callback.answer("Mahsulot topilmadi", show_alert=True)
        return
    await state.update_data(discount_product_id=product_id)
    await state.set_state(SetDiscount.price)
    await callback.answer()
    await callback.message.answer(
        f"<b>{product['name']}</b> uchun chegirma narxini kiriting (so'mda):",
        parse_mode="HTML",
    )


@router.message(SetDiscount.price)
async def product_discount_price_input(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None or not await is_owner_level(message.from_user.id):
        await state.clear()
        return
    try:
        price = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 25000")
        return
    if price <= 0:
        await message.answer("Narx 0 dan katta bo'lishi kerak. Qaytadan kiriting:")
        return

    data = await state.get_data()
    product = await db.get_product(shop_id, data["discount_product_id"])
    if not product:
        await message.answer("❌ Mahsulot topilmadi.")
        await state.clear()
        return
    if price < product["price"]:
        await message.answer(
            f"❌ Chegirma narxi tannarxdan ({product['price']:.0f} so'm) past bo'lishi mumkin emas, "
            f"aks holda zararga ketasiz. Qaytadan kiriting:"
        )
        return

    await state.update_data(discount_price=price)
    await state.set_state(SetDiscount.days)
    await message.answer("Chegirma necha kun amal qiladi? Masalan: 7")


@router.message(SetDiscount.days)
async def product_discount_days_input(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None or not await is_owner_level(message.from_user.id):
        await state.clear()
        return
    try:
        days = int(message.text.replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat butun son kiriting. Masalan: 7")
        return
    if days <= 0:
        await message.answer("Kunlar soni 0 dan katta bo'lishi kerak. Qaytadan kiriting:")
        return

    data = await state.get_data()
    product = await db.get_product(shop_id, data["discount_product_id"])
    if not product:
        await message.answer("❌ Mahsulot topilmadi.")
        await state.clear()
        return

    await db.set_product_discount(shop_id, product["id"], data["discount_price"], days)
    await state.clear()
    await message.answer(
        f"✅ <b>{product['name']}</b> uchun {data['discount_price']:.0f} so'mlik chegirma "
        f"{days} kunga belgilandi.",
        parse_mode="HTML",
    )


# ---------- MAHSULOTNI BO'LIMLAR ORASIDA KO'CHIRISH / BO'LIMDAN CHIQARISH ----------

@router.callback_query(F.data.startswith("prod_move_") & ~F.data.startswith("prod_move_to_"))
async def product_move_start_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    product_id = int(callback.data.split("_")[-1])
    product = await db.get_product(shop_id, product_id)
    if not product:
        await callback.answer("Mahsulot topilmadi", show_alert=True)
        return
    categories = await db.get_categories(shop_id)
    if not categories:
        await callback.answer("Hozircha bo'lim yo'q. Avval bo'lim yarating.", show_alert=True)
        return
    await callback.answer()
    await callback.message.answer(
        f"<b>{product['name']}</b> qaysi bo'limga ko'chirilsin?",
        reply_markup=kb.category_move_kb(categories, product_id, product.get("category_id")),
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("prod_move_to_"))
async def product_move_to_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    parts = callback.data.split("_")
    product_id, category_id = int(parts[-2]), int(parts[-1])
    category = await db.get_category(shop_id, category_id)
    if not category:
        await callback.answer("Bo'lim topilmadi", show_alert=True)
        return
    moved = await db.set_product_category(shop_id, product_id, category_id)
    if not moved:
        await callback.answer("Mahsulot topilmadi", show_alert=True)
        return
    await callback.answer(f"✅ \"{category['name']}\" bo'limiga ko'chirildi", show_alert=True)
    try:
        await callback.message.delete()
    except Exception:
        pass


@router.callback_query(F.data.startswith("prod_unassign_"))
async def product_unassign_cb(callback: CallbackQuery):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    product_id = int(callback.data.split("_")[-1])
    removed = await db.set_product_category(shop_id, product_id, None)
    if not removed:
        await callback.answer("Mahsulot topilmadi", show_alert=True)
        return
    await callback.answer("✅ Mahsulot bo'limdan chiqarildi", show_alert=True)
    try:
        await callback.message.delete()
    except Exception:
        pass


# ---------- MIQDORNI O'ZGARTIRISH ----------

async def _channel_caption(product: dict) -> str:
    return f"📦 {product['name']} | {product['price']:.0f} so'm | {product['quantity']:.0f} dona"



# ---------- SEKIN SOTILADIGAN TOVARLAR ----------

@router.message(F.text == "🐌 Sekin sotiladigan tovarlar")
async def stale_products(message: Message):
    shop_id = await _require_shop(message)
    if shop_id is None:
        return
    stale = await db.get_stale_products(shop_id, days=30, limit=10)
    if not stale:
        await message.answer("✅ Hozircha 30 kundan beri sotilmagan tovar yo'q.")
        return

    text = (
        "🐌 <b>Top 10 - eng sekin sotiladigan tovarlar</b>\n"
        "(30 kundan ortiq sotilmagan, pul band bo'lib turibdi)\n\n"
    )
    for i, p in enumerate(stale, 1):
        reference = p["reference_date"][:10]
        text += f"{i}. {p['name']} — {p['quantity']:.0f} dona qoldi (oxirgi harakat: {reference})\n"
    text += "\n💡 Bularga chegirma qilib tezroq sotib, pulni aylantirish tavsiya etiladi."

    await message.answer(text, parse_mode="HTML")


# ---------- AI BUYURTMA TAVSIYASI - 16-BOSQICH ----------

@router.message(F.text == "🤖 AI buyurtma tavsiyasi")
async def ai_restock_suggestions(message: Message):
    """AI BUYURTMA TAVSIYASI - 16-BOSQICH: "🧾 Olinishi kerak bo'lgan
    tovarlar"dan farqi - u yerda faqat xodim QO'LDA belgilagan chegara
    (alert_quantity) tekshiriladi, bu yerda esa HAR BIR mahsulotning
    haqiqiy SOTILISH TEZLIGI (oxirgi 30 kun) joriy QOLDIQQA taqqoslanib,
    "necha kunlik zaxira qolgani" avtomatik hisoblanadi va yetkazib berish
    muddatidan kam bo'lsa - shu zahoti tavsiya sifatida ko'rsatiladi."""
    shop_id = await _require_shop(message)
    if shop_id is None:
        return

    lead_time = await db.get_restock_lead_time_days(shop_id)
    suggestions = await db.get_ai_restock_suggestions(shop_id, lookback_days=30, lead_time_days=lead_time)

    if not suggestions:
        await message.answer(
            "🤖 <b>AI buyurtma tavsiyasi</b>\n\n"
            "✅ Hozircha hech qanday mahsulot uchun shoshilinch buyurtma kerak emas "
            f"(hisoblash oxirgi 30 kunlik sotuvlar va {lead_time} kunlik yetkazib berish "
            "muddati asosida qilinadi).",
            parse_mode="HTML",
        )
        return

    lines = [
        "🤖 <b>AI buyurtma tavsiyasi</b>\n"
        f"(oxirgi 30 kunlik sotilish tezligi + {lead_time} kunlik yetkazib berish "
        "muddati asosida, eng shoshilinchidan boshlab)\n",
    ]
    for s in suggestions:
        p = s["product"]
        urgency = "🔴" if s["days_left"] <= 0 else ("🟠" if s["days_left"] <= lead_time / 2 else "🟡")
        lines.append(
            f"{urgency} <b>{p['name']}</b>\n"
            f"   Qoldiq: {p['quantity']:.0f} dona | Kuniga ~{s['daily_sales_rate']:.1f} dona sotiladi\n"
            f"   Taxminan {max(s['days_left'], 0):.0f} kunga yetadi\n"
            f"   💡 Tavsiya: kamida {s['suggested_qty']:.0f} dona buyurtma bering"
        )

    await message.answer("\n\n".join(lines), parse_mode="HTML")


# ---------- OLINISHI KERAK BO'LGAN TOVARLAR ----------

async def _send_restock_list(message: Message, shop_id: int, manage: bool = True):
    low_stock = await db.get_low_stock_products(shop_id)
    manual_items = await db.get_manual_restock_items(shop_id)

    if not low_stock and not manual_items:
        await message.answer(
            "✅ Hozircha olinishi kerak bo'lgan tovar yo'q.\n"
            "Kerak bo'lsa, pastdagi tugma orqali qo'lda qo'shishingiz mumkin.",
            reply_markup=kb.restock_kb(low_stock, manual_items, manage=manage),
        )
        return

    text = "🧾 <b>Olinishi kerak bo'lgan tovarlar</b>\n\n"
    if low_stock:
        text += "📉 <b>Skladda kamayib qolgan (avtomatik):</b>\n"
        for p in low_stock:
            text += (
                f"• {p['name']} — {p['quantity']:.0f} dona qoldi "
                f"(chegara: {p['alert_quantity']:.0f})\n"
            )
        text += "\n"
    if manual_items:
        text += "✍️ <b>Qo'lda qo'shilgan:</b>\n"
        for item in manual_items:
            note = f" — {item['note']}" if item.get("note") else ""
            text += f"• {item['name']}{note}\n"

    await message.answer(
        text, reply_markup=kb.restock_kb(low_stock, manual_items, manage=manage), parse_mode="HTML"
    )


@router.message(F.text == "🧾 Olinishi kerak bo'lgan tovarlar")
async def restock_list(message: Message, state: FSMContext):
    await state.clear()
    shop_id = await _require_shop(message)
    if shop_id is None:
        return
    # Faqat do'kon egasi tovarni "sotib olindi" deb belgilay oladi - sotuvchi
    # bu bo'limni faqat ko'rish uchun ochadi.
    manage = await db.is_owner(message.from_user.id)
    await _send_restock_list(message, shop_id, manage=manage)


@router.callback_query(F.data == "restock_add")
async def restock_add_start(callback: CallbackQuery, state: FSMContext):
    if await _require_shop_cb(callback) is None:
        return
    await state.set_state(AddRestockItem.name)
    await callback.message.answer("Olinishi kerak bo'lgan tovar nomini kiriting:")
    await callback.answer()


@router.message(AddRestockItem.name)
async def restock_add_name(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await state.clear()
        return

    name = message.text.strip()
    existing = await db.find_product_by_name(shop_id, name)
    if existing:
        await message.answer(
            f"❗ Bunday mahsulot skladda mavjud: <b>{existing['name']}</b> "
            f"(hozir {existing['quantity']:.0f} dona bor).\n"
            f"Boshqa nom kiriting, yoki agar shu mahsulotdan yana kerak bo'lsa, "
            f"\"➕ Mahsulot qo'shish\" orqali unga miqdor qo'shing.",
            parse_mode="HTML",
        )
        return

    await state.update_data(name=name)
    await state.set_state(AddRestockItem.note)
    await message.answer("Izoh qoldirasizmi? (kerak bo'lmasa \"-\" deb yozing)")


@router.message(AddRestockItem.note)
async def restock_add_note(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await state.clear()
        return
    note = message.text.strip()
    if note == "-":
        note = None
    data = await state.get_data()
    await db.add_manual_restock_item(shop_id, data["name"], note)
    await state.clear()
    await message.answer(
        f"✅ \"{data['name']}\" olinishi kerak bo'lgan tovarlar ro'yxatiga qo'shildi."
    )
    manage = await db.is_owner(message.from_user.id)
    await _send_restock_list(message, shop_id, manage=manage)


@router.callback_query(F.data.startswith("lowstock_notbought_"))
async def lowstock_not_bought_cb(callback: CallbackQuery):
    await callback.answer("Belgilandi, tovar ro'yxatda qolaveradi.")


@router.callback_query(F.data.startswith("lowstock_bought_"))
async def lowstock_bought_cb(callback: CallbackQuery, state: FSMContext):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    product_id = int(callback.data.split("_")[-1])
    product = await db.get_product(shop_id, product_id)
    if not product:
        await callback.answer("Mahsulot topilmadi", show_alert=True)
        return

    await state.set_state(RestockPurchase.quantity)
    await state.update_data(restock_type="product", product_id=product_id, name=product["name"])
    await callback.message.answer(
        f"<b>{product['name']}</b> — necha dona sotib olindi?", parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data.startswith("restock_done_"))
async def restock_done_cb(callback: CallbackQuery, state: FSMContext):
    shop_id = await _require_shop_cb(callback)
    if shop_id is None:
        return
    item_id = int(callback.data.split("_")[-1])
    item = await db.get_manual_restock_item(shop_id, item_id)
    if not item:
        await callback.answer("Topilmadi", show_alert=True)
        return

    await state.set_state(RestockPurchase.quantity)
    await state.update_data(restock_type="manual", manual_id=item_id, name=item["name"])
    await callback.message.answer(
        f"<b>{item['name']}</b> — necha dona sotib olindi?", parse_mode="HTML"
    )
    await callback.answer()


# ---------- SOTIB OLINGAN TOVARNI SKLADGA AVTOMATIK QO'SHISH ----------

@router.message(RestockPurchase.quantity)
async def restock_purchase_quantity(message: Message, state: FSMContext):
    try:
        quantity = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 10")
        return
    if quantity <= 0:
        await message.answer("Miqdor 0 dan katta bo'lishi kerak. Qaytadan kiriting:")
        return

    await state.update_data(quantity=quantity)
    await state.set_state(RestockPurchase.price)
    await message.answer("Qanchaga olindi (tannarxi, faqat raqam)?")


@router.message(RestockPurchase.price)
async def restock_purchase_price(message: Message, state: FSMContext):
    try:
        price = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 25000")
        return

    await state.update_data(price=price)
    await state.set_state(RestockPurchase.sell_price)
    await message.answer("Qanchaga sotasiz (savdo narxi)?")


@router.message(RestockPurchase.sell_price)
async def restock_purchase_sell_price(message: Message, state: FSMContext):
    try:
        sell_price = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 30000")
        return

    data = await state.get_data()
    if sell_price < data["price"]:
        await message.answer(
            f"❌ Savdo narxi tannarxdan ({data['price']:.0f} so'm) past bo'lishi mumkin emas. "
            f"Qaytadan kiriting:"
        )
        return

    await state.update_data(sell_price=sell_price)
    await state.set_state(RestockPurchase.min_price)
    await message.answer("Qanchagacha tushirib berish mumkin (eng past narx)?")


@router.message(RestockPurchase.min_price)
async def restock_purchase_min_price(message: Message, state: FSMContext):
    try:
        min_price = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 27000")
        return

    data = await state.get_data()
    if min_price > data["sell_price"]:
        await message.answer(
            f"Eng past narx savdo narxidan ({data['sell_price']:.0f} so'm) katta bo'lmasligi kerak. "
            f"Qaytadan kiriting:"
        )
        return
    if min_price < data["price"]:
        await message.answer(
            f"❌ Eng past narx tannarxdan ({data['price']:.0f} so'm) past bo'lishi mumkin emas. "
            f"Qaytadan kiriting:"
        )
        return

    await state.update_data(min_price=min_price)
    await state.set_state(RestockPurchase.alert_quantity)
    await message.answer(
        "Mahsulot nechta qolganda ogohlantirish yuborilsin?\n"
        "(Kerak bo'lmasa 0 deb yozing)"
    )


@router.message(RestockPurchase.alert_quantity)
async def restock_purchase_alert_quantity(message: Message, state: FSMContext):
    try:
        alert_quantity = float(message.text.replace(",", ".").replace(" ", ""))
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting. Masalan: 5 (kerak bo'lmasa 0)")
        return
    await state.update_data(alert_quantity=alert_quantity if alert_quantity > 0 else None)
    await _finalize_restock_purchase(message, state)


async def _finalize_restock_purchase(message: Message, state: FSMContext):
    shop_id = await get_shop_id(message.from_user.id)
    if shop_id is None:
        await state.clear()
        return
    data = await state.get_data()
    alert_quantity = data.get("alert_quantity")
    alert_line = f"Ogohlantirish chegarasi: {alert_quantity:.0f} dona\n" if alert_quantity else ""

    if data["restock_type"] == "product":
        result = await db.update_product_purchase(
            shop_id, data["product_id"], data["quantity"], data["price"], data["sell_price"],
            data["min_price"], alert_quantity=alert_quantity,
        )
        if result is None:
            await state.clear()
            await message.answer("❌ Mahsulot topilmadi.")
            return
        new_quantity, weighted_price = result
        product = await db.get_product(shop_id, data["product_id"])
        if product and product.get("channel_message_id") and config.CHANNEL_ID:
            try:
                await message.bot.edit_message_caption(
                    chat_id=config.CHANNEL_ID,
                    message_id=product["channel_message_id"],
                    caption=await _channel_caption(product),
                )
            except Exception as e:
                logging.warning(f"Kanaldagi postni yangilab bo'lmadi: {e}")

        await state.clear()
        await message.answer(
            f"✅ <b>{data['name']}</b> skladga qo'shildi.\n"
            f"Qo'shildi: {data['quantity']:.0f} dona ({data['price']:.0f} so'mdan)\n"
            f"Yangi umumiy miqdor: {new_quantity:.0f} dona\n"
            f"O'rtacha tannarx: {weighted_price:.0f} so'm\nSavdo narxi: {data['sell_price']:.0f} so'm\n"
            f"Eng past narx: {data['min_price']:.0f} so'm\n{alert_line}",
            parse_mode="HTML"
        )
        await _send_restock_list(message, shop_id)
    else:
        await db.add_product(
            shop_id, data["name"], data["price"], data["quantity"], None,
            sell_price=data["sell_price"], min_price=data["min_price"],
            alert_quantity=alert_quantity,
        )
        await db.delete_manual_restock_item(shop_id, data["manual_id"])
        await state.clear()
        await message.answer(
            f"✅ <b>{data['name']}</b> yangi mahsulot sifatida skladga qo'shildi.\n"
            f"Miqdori: {data['quantity']:.0f} dona\nTannarx: {data['price']:.0f} so'm\n"
            f"Savdo narxi: {data['sell_price']:.0f} so'm\nEng past narx: {data['min_price']:.0f} so'm\n"
            f"{alert_line}",
            parse_mode="HTML"
        )
        await _send_restock_list(message, shop_id)
